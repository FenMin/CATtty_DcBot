import discord
from discord.ext import commands
from openpyxl import workbook, load_workbook
import json

with open('./data/data.json' , 'r') as f:
    data = json.load(f)

path = data['record_xlsx']

wb = load_workbook(path)
ws = wb.active
count = 0
value = 0
def check_A_list_len():
    x=1
    while ws[f'A{x}'].value != None:
        x+=1

    if x == 1:
        return 1   #no data
    elif x == 2:
        return 1   #only 1 data
    else:
        return int(x-1)  #len = minute "start num"

def add_count(id):
    for i in range(1 , check_A_list_len()+1):  #list of space we use
        if ws[f'A1'].value == None:
            ws[f'A{i}'].value = id
            ws[f'B{i}'].value = 1
            value = ws[f'B1'].value
            break
        
        elif i == check_A_list_len() and ws['A1'].value != None and id != ws[f'A{i}'].value:
            ws[f'A{i+1}'].value = id
            ws[f'B{i+1}'].value = 1 
            value = ws[f'B{i+1}'].value
            break

        elif id == ws[f'A{i}'].value:
            ws[f'B{i}'].value+=1 
            value = ws[f'B{i}'].value
            break

    wb.save(path)
    return value
#---------------------------------------------------------cog start

class record(commands.Cog):

    def __init__(self, bot):
        self.bot = bot

    @commands.Cog.listener()
    async def on_message(self, msg):
        channel = msg.channel
        keyw = "2202"
        if msg.content == keyw:
            id = msg.author.id
            count = add_count(int(id/100000000))
            await channel.send(f'*{msg.author.name}* 你已經{keyw}了 ***{count}*** 次')
            

    
                
        




def setup(bot):
    bot.add_cog(record(bot))